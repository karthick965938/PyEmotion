import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#Location of original excel sheet
fileLocation ='data.xlsx'

#Location of new file which can be the same as original file
writeLocation='dataNew.xlsx'

data = {'Name':['Tom','Paul','Jeremy'],'Age':[32,43,34],'Salary':[20000,34000,32000]}

#The dataframe you want to add
df = pd.DataFrame(data)

#Load existing sheet as it is
book = load_workbook(fileLocation)
#create a new sheet
sheet = book.create_sheet("Sheet Name")

#Load dataframe into new sheet
for row in dataframe_to_rows(df, index=False, header=True):
    sheet.append(row)

#Save the modified excel at desired location    
book.save(writeLocation)